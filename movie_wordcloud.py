#Selenium模块是一个自动化测试工具，能够驱动浏览器模拟人的操作，如单击、键盘输入等。
import re
import jieba
from jieba.analyse import *
from openpyxl import load_workbook
from wordcloud import WordCloud
import matplotlib.pyplot as plt

#读取excel文件工作表Sheet1
workbook = load_workbook('豆瓣电影Top250.xlsx')
worksheet = workbook['Sheet1']
#将Sheet1中F列内容写入txt文件
data = []
f = open('电影信息.txt', 'w')
for row in range(2, worksheet.max_row + 1):
    data.append(worksheet['I'+str(row)].value)
f.write(str(data))
f.close()


#生成词云
#读取txt
f = open('电影信息.txt', 'r', encoding='gbk').read()
#分词
sep_list = jieba.lcut(f)
print('分词', type(sep_list), len(sep_list), sep_list)
#过滤停用词
# stopwords为停用词list，stop.txt停用词一行一个
stopwords = [line.strip() for line in open('stop.txt', 'r', encoding='utf-8').readlines()]
print('停用词', type(stopwords), stopwords)
outstr = ''
for word in sep_list:
    if word not in stopwords:
        outstr += word
print('过滤停用词后', type(outstr), outstr)
#过滤后重新分词

outstr = jieba.lcut(outstr)
outstr = ' '.join(outstr)
print('再次分词后', type(outstr), outstr)
#生成词云图
#设置词云使用的字体
font = r'C:\Windows\Fonts\simsun.ttc'
wc = WordCloud(font_path=font, width=2400, height=1200, max_words=100)
wc.generate(outstr)
wc.to_file('词云.jpg')
plt.figure(dpi=100)
plt.imshow(wc, interpolation='catrom')
plt.axis('off')
plt.show()
plt.close()

#生成词频
for keyword, weight in extract_tags(outstr, topK=50, withWeight=True, allowPOS=()):
    print('%s %s' % (keyword, weight))