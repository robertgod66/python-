import numpy as np
from openpyxl import load_workbook
import pandas
from spsspro.algorithm import supervised_learning
workbook = load_workbook('版本2_缺失值处理_豆瓣电影Top250.xlsx')
worksheet = workbook['编码数据_版本2_缺失值处理']

# 把数据写入 xi，y 变量中
x = [[] for i in range(5)]
for row in range(2, worksheet.max_row + 1):
  for name in ['D', 'E', 'F', 'G', 'H']:
     x[ord(name) -ord('D')].append(worksheet[name + str(row)].value)
y = []
for row in range(2, worksheet.max_row + 1):
    y.append(worksheet['C' + str(row)].value)
for i in range(len(x)):
  x[i] = np.array(x[i])
y = np.array(y)
#ExtraTrees回归，输入参数详细可以光标放置函数括号内按shift+tab查看，输出结果参考spsspro模板分析报告
result = supervised_learning.extra_tree_regression(data_x=x, data_y=y)
print(result)