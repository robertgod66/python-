import numpy as np
from scipy import optimize
from openpyxl import load_workbook

workbook = load_workbook('版本2_缺失值处理_豆瓣电影Top250.xlsx')
worksheet = workbook['编码数据_版本2_缺失值处理']

# 把数据写入 xi，y 变量中
x = [[] for i in range(5)]
for row in range(2, worksheet.max_row + 1):
  for name in ['D', 'E', 'F', 'G', 'H']:
     x[ord(name) -ord('D')].append(worksheet[name + str(row)].value)
y = []
for row in range(2, worksheet.max_row + 1):
    y.append(worksheet['C' + str(row)].value)
for i in range(len(x)):
  x[i] = np.array(x[i])
y = np.array(y)

def residuals(p):
  k0, k1, k2, k3, k4,  b = p
  return y -(k0 * x[0] + k1 * x[1] + k2 * x[2] + k3 * x[3] + k4 * x[4] + b)
r = optimize.leastsq(residuals, [1, 1, 1, 1, 1, 0])

# 获取拟合得到的参数
params = r[0]

# 根据拟合得到的参数进行预测
k0, k1, k2, k3, k4, b = params
y_pred = k0 * x[0] + k1 * x[1] + k2 * x[2] + k3 * x[3] + k4 * x[4] + b

# 计算相关系数矩阵
correlation_matrix = np.corrcoef(y, y_pred)
# 提取相关系数
correlation_coefficient = correlation_matrix[0, 1]

print("拟合得到的参数:", params)
print("线性模型的相关系数:", correlation_coefficient)


